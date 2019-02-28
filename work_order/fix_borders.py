from openpyxl.worksheet import Worksheet
from openpyxl.reader.worksheet import WorkSheetParser
from openpyxl.worksheet.merge import MergeCells
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import range_boundaries


def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
    """

    # apply patch 1

    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        cr = CellRange(range_string=range_string, min_col=start_column, min_row=start_row,
                       max_col=end_column, max_row=end_row)
        self.merged_cells.add(cr.coord)
        # self._clean_merge_range(cr)

    Worksheet.merge_cells = merge_cells

    # apply patch 2

    def parse_merge(self, element):
        merged = MergeCells.from_tree(element)
        self.ws.merged_cells.ranges = merged.mergeCell
        # for cr in merged.mergeCell:
        #     self.ws._clean_merge_range(cr)

    WorkSheetParser.parse_merge = parse_merge


patch_worksheet()
